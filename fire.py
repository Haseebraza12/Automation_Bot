import time
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import re
import concurrent.futures
from tqdm import tqdm
import threading
import tkinter as tk
from tkinter import ttk,messagebox,filedialog
import queue
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import csv
from datetime import datetime
from openpyxl.utils import get_column_letter
# Sample form_data (this should be replaced with actual form data as required)
'''form_data = {
    "date":"13/12/2024",
    "time":"7:25 pm",
    "person represented":"no",
    "case number":"12",
  "name": "John Doe",
  "first name": "Joe",
    "last name":"Doe",
    "phone": "123-456-7890",
    "p1": "123",
    "p2":"456",
    "p3":"7890",
    "email": "dummyemail@example.com",
    "address": "123 Main Street, Suite 100",
    "city": "SampleCity",
    "state": "SampleState",
    "zip": "12345",
    "company":"no",
    "case":"no",
    "unit number":"12",
    "country":"USA",
    "message": """My name is John Doe, and I am requesting records from the Code Enforcement Office.

In accordance with the [Insert Local Open Records Act], I am requesting a list of all open code violations for residential properties over the past 30 days. Specifically, I am interested in violations related to damaged or decayed roofs, mold, broken windows, boarded-up windows and doors, overgrown weeds and grass, trash and debris, rodent infestations or unsanitary conditions, flaking or peeling paint, vacant and unsecured structures, and any buildings deemed dangerous, uninhabitable, or unfit for occupancy.

Please provide the details of these violations, including the nature of the violation, the address of the property, and the date of the violation. If possible, I would appreciate receiving the information in a digital format, such as a .csv file or a searchable PDF. However, I am happy to accept the information in any format that is convenient for your office.

Thank you for your assistance in this matter.

Sincerely,
John Doe"""
}'''

def create_driver():
    # Set Chrome options
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless=new")  # Use the latest headless mode
    chrome_options.add_argument("--no-sandbox")  # Bypass OS security model
    chrome_options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems
    chrome_options.add_argument("--window-size=1920x1080")  # Set window size to avoid detection issues
    
    # Optimize performance
    chrome_options.add_argument("--disable-extensions")  # Disable extensions to speed up
    chrome_options.add_argument("--disable-gpu")  # Disable GPU for headless mode
    chrome_options.add_argument("--disable-software-rasterizer")  # Avoid software rendering
    chrome_options.add_argument("--blink-settings=imagesEnabled=false")  # Disable loading images
    chrome_options.add_argument("--enable-automation")  # Indicate automation use
    chrome_options.add_argument("--disable-infobars")  # Disable 'Chrome is being controlled' message
    
    # Additional security & detection prevention
    chrome_options.add_argument("--disable-popup-blocking")  # Prevent blocking popups
    chrome_options.add_argument("--incognito")  # Run in incognito mode for cleaner sessions
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.199 Safari/537.36")  # Set user agent to reduce detection

    # Enable logging for debugging purposes (optional)
    chrome_options.add_argument("--log-level=3")  # Minimize logging output (DEBUG=0, INFO=1, WARNING=2, ERROR=3)
    
    # Return the initialized driver
    return webdriver.Chrome(options=chrome_options)

def fill_field(driver, wait, xpath, value):
    try:
        field = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        field.clear()
        field.send_keys(value)
        print(f"Filled field at {xpath} with value: {value}")
    except Exception as e:
        print(f"Error filling field at {xpath}: {str(e)}")

def handle_cantonga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_maconbibbcountyga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["name"],       
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/input": form_data["zip"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["phone"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[13]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[14]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}


def handle_woodstockga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("View Only")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_smyrnaga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["first name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["last name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[15]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[14]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review/inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        
        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_austellga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(60)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_albanyga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[15]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Handle checkboxes
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[17]/div/div[1]/div/div/div/div/div/div/div"
        checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox1_xpath)))
        driver.execute_script("arguments[0].click();", checkbox)
        print(f"Checked {checkbox1_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}


def handle_forsythcountyga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[13]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[14]/div/div[1]/div/div/div/div/div/div/div"
        checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox1_xpath)))
        driver.execute_script("arguments[0].click();", checkbox)
        print(f"Checked {checkbox1_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            if security_key_match:
                security_key = security_key_match.group(1)
            else:
                security_key = ""
                print("Security key not found in confirmation message")
            if reference_number_match:
                reference_number = reference_number_match.group(1)
            else:
                reference_number = ""
                print("Reference number not found in confirmation message")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}

def handle_fairburnga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/div/div/div/div/div/div/div"
        for checkbox_xpath in [checkbox1_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(40)

        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}

def handle_collegeparkga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}
    except Exception as e:
            print(f"Error in form handling: {str(e)}")
            driver.save_screenshot("error_main.png")
            return {"status": "Failed", "error": str(e)}
    
def handle_conyersga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div[1]/div/div/div/div/div/div/div"
        for checkbox_xpath in [checkbox1_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(30)

        # Delay of 15 seconds after form submission
        time.sleep(40)

        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}


def handle_spaldingcountyga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(40)

        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}

   

def handle_cityofaugustaga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
        
            
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[1]/div/input": form_data["email"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[2]/div/input": form_data["name"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[3]/div/input": form_data["phone"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[4]/div/textarea": form_data["address"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[1]/div/div/input": form_data["city"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[3]/div/div/input": form_data["zip"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[6]/div/input": form_data["company"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[1]/div[2]/div[1]/div[2]/div/div[1]"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[2]/div/div/div/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Alabama")  # Adjust text as per available options
        print("Selected dropdown option")
#checkbox
        checkbox1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[3]/div[2]/div/div/input"
        for checkbox_xpath in [checkbox1_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[4]/button"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_cityoffayettevillega_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[1]/div/input": form_data["email"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[2]/div/input": form_data["name"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[3]/div/input": form_data["phone"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[4]/div/textarea": form_data["address"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[1]/div/div/input": form_data["city"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[3]/div/div/input": form_data["zip"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[6]/div/input": form_data["company"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[1]/div[2]/div[1]/div[2]/div/div[1]"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle department dropdown/autocomplete
        dropdown1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[1]/div[4]/div/div/div/div/div[1]/input"
        dropdown1 = wait.until(EC.element_to_be_clickable((By.XPATH, dropdown1_xpath)))
        
        # Clear existing value
        dropdown1.clear()
        
        # Click to focus
        dropdown1.click()
        time.sleep(1)
        
        # Type value
        dropdown1.send_keys("City Hall")
        time.sleep(1)
        
        # Press down arrow and enter to select from autocomplete
        actions = ActionChains(driver)
        actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()
        time.sleep(1)
        print("Entered 'City Hall' in department field")

        # Handle state dropdown
        dropdown2_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[2]/div/div/div/select"
        dropdown2 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown2_xpath))))
        dropdown2.select_by_visible_text("Alabama")
        print("Selected state dropdown option")

        # Handle checkbox
        checkbox1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[3]/div[2]/div/div/input"
        checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox1_xpath)))
        driver.execute_script("arguments[0].click();", checkbox)
        print("Checked agreement checkbox")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[4]/button"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_unioncityga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")
        
        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[13]/div[1]/div/div/div/div/div/div/div"
        checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox1_xpath)))
        driver.execute_script("arguments[0].click();", checkbox)
        print(f"Checked {checkbox1_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 30 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_hapevillega_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")
        
        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div[1]/div/div/div/div/div/div/div"
        checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox1_xpath)))
        driver.execute_script("arguments[0].click();", checkbox)
        print(f"Checked {checkbox1_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_eastpointga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[12]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[14]/div/div[1]/div/div/div/div/div/div/div"
        for checkbox_xpath in [checkbox1_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}



def handle_sandyspringsga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["zip"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[14]/div/div[1]/input": form_data["case"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[16]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("No")  # Adjust text as per available options
        print("Selected dropdown option")

        dropdown2_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[15]/div/div[1]/select"
        dropdown2 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown2_xpath))))
        available_options = [o.text.strip() for o in dropdown2.options]
        print("Available options in dropdown:", available_options)
        dropdown2.select_by_visible_text("N/A - I am not requesting an accident report")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 30 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Request number:')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Security key: (\S+)", confirmation_message)
            reference_number_match = re.search(r"Request number: (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_roswellga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div[1]/div/div/div/div/div/div/div"
        checkbox2_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[19]/div[1]/div/div/div/div/div/div/div"
        checkbox3_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[21]/div[1]/div/div/div/div/div/div/div"
        checkbox4_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[22]/div[1]/div/div/div/div/div/div/div"
        for checkbox_xpath in [checkbox1_xpath, checkbox2_xpath, checkbox3_xpath, checkbox4_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 30 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Request Number:')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Security Key:\s+(\S+)", confirmation_message)
            reference_number_match = re.search(r"Request Number:\s+(\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_alpharettaga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[8]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[20]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[9]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[23]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[24]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[25]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[26]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[7]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle checkbox selection
        checkbox1_xpath = "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[19]/div[1]/div/div/div/div/div/div/div"
        checkbox2_xpath = "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[10]/div[1]/div/div/div/div/div/div/div"
        checkbox3_xpath = "/html/body/div[1]/div[2]/main/div/div[1]/form/div[2]/div/div[11]/div[1]/div/div/div/div/div/div/div"
        for checkbox_xpath in [checkbox1_xpath, checkbox2_xpath, checkbox3_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/main/div/div[1]/form/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 30 seconds after form submission
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}

def handle_norcrossga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Other")  # Adjust text as per available options
        print("Selected dropdown option")

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["case number"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 30 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_cityofgriffin_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[1]/vi-form-field-edit/vi-field-fullname-edit/div/span[1]/input": form_data["first name"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[1]/vi-form-field-edit/vi-field-fullname-edit/div/span[2]/input": form_data["last name"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[4]/vi-form-field-edit/vi-field-email-edit/div/input": form_data["email"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[3]/vi-form-field-edit/vi-field-phone-edit/div/span/input": form_data["phone"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/div[2]/ol/li[2]/input": form_data["email"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[1]/input": form_data["address"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[2]/input": form_data["unit number"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[3]/span[1]/input": form_data["city"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[3]/span[2]/input": form_data["state"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[4]/input": form_data["zip"],
            "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[2]/vi-form-field-edit/vi-field-fulladdress-edit/div/span[5]/input": form_data["country"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[5]/vi-form-field-edit/vi-field-paragraphtext-edit/div/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/div/form-builder-submit-pagination/div/div[6]/vi-form-field-edit/vi-field-singledropdown-edit/div/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Yes")  # Adjust text as per available options
        print("Selected dropdown option")
      
        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[2]/div/div[2]/div[3]/div[2]/section/div/form/div/form-builder-submit-actions/div/a[2]"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    


def handle_henrycounty_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[3]/div/div/div[2]/div/input": form_data["first name"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[4]/div/div/div[2]/div/input": form_data["last name"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[6]/div/div/div[2]/div/input": form_data["email"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[11]/div/div/div[2]/div/input": form_data["phone"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[7]/div/div/div[2]/div/input": form_data["address"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[8]/div/div/div[2]/div/input": form_data["city"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[9]/div/div/div[2]/div/input": form_data["state"],
            "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[10]/div/div/div[2]/div/input": form_data["zip"],
          
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[3]/div/div/div/div[13]/div/div/div[2]/textarea[1]"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div/ob-bottom-navigation/div/div[1]/section/div/div/form/div/div[4]/button[2]/span/span"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    


def handle_fayettecountyga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[1]": form_data["first name"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[2]": form_data["last name"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[5]": form_data["email"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[4]": form_data["phone"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[6]": form_data["email"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[7]": form_data["address"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[8]": form_data["city"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[9]": form_data["state"],
            "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/input[10]": form_data["zip"],
          
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div/div[1]/div/div[5]/div/div[1]/div[1]/form/label[12]/input"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_formsite_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/form/div[2]/div[6]/input": form_data["date"],
            "/html/body/form/div[2]/div[9]/input": form_data["name"],
            "/html/body/form/div[2]/div[22]/input": form_data["phone"],
            "/html/body/form/div[2]/div[21]/input": form_data["email"],
            "/html/body/form/div[2]/div[14]/input": form_data["address"],
            "//html/body/form/div[2]/div[17]/input": form_data["city"],
            "/html/body/form/div[2]/div[19]/input": form_data["zip"],
            "/html/body/form/div[2]/div[30]/input": form_data["date"],
            "/html/body/form/div[2]/div[40]/input": form_data["address"],
            "/html/body/form/div[2]/div[42]/input": form_data["city"],
            "/html/body/form/div[2]/div[44]/input": form_data["zip"],
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/form/div[2]/div[26]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle first dropdown
        try:
            dropdown0_xpath = "/html/body/form/div[2]/div[24]/select"
            dropdown0 = wait.until(EC.presence_of_element_located((By.XPATH, dropdown0_xpath)))
            select0 = Select(dropdown0)
            
            # Wait for options to load
            time.sleep(2)
            
            # Get available options
            options = [o for o in select0.options if o.text.strip()]
            if options:
                select0.select_by_visible_text(options[-1].text)  # Select last non-empty option
                print(f"Selected {options[-1].text} from first dropdown")
            else:
                print("No valid options found in first dropdown")
                
        except Exception as e:
            print(f"Error handling first dropdown: {str(e)}")
            driver.save_screenshot("dropdown0_error.png")

        # Handle state dropdowns
        for dropdown_xpath in ["/html/body/form/div[2]/div[18]/select", 
                             "/html/body/form/div[2]/div[43]/select"]:
            try:
                dropdown = wait.until(EC.presence_of_element_located((By.XPATH, dropdown_xpath)))
                select = Select(dropdown)
                
                # Wait for options to load
                time.sleep(2)
                
                # Try multiple selection methods
                try:
                    select.select_by_visible_text("Alabama")
                except:
                    try:
                        select.select_by_value("AL")
                    except:
                        try:
                            # Get all non-empty options
                            options = [o for o in select.options if o.text.strip()]
                            if options:
                                select.select_by_visible_text(options[1].text)  # Select first non-empty option
                            else:
                                print(f"No valid options found in dropdown {dropdown_xpath}")
                        except:
                            print(f"Could not select state in dropdown {dropdown_xpath}")
                            
                print(f"Handled state selection for {dropdown_xpath}")
                
            except Exception as e:
                print(f"Error handling state dropdown {dropdown_xpath}: {str(e)}")
                driver.save_screenshot(f"dropdown_error_{dropdown_xpath.replace('/', '_')}.png")

        # Handle checkboxes
        checkbox_xpaths = [
            "/html/body/form/div[2]/div[7]/table/tbody/tr[2]/td/label",
            "/html/body/form/div[2]/div[31]/table/tbody/tr[1]/td/label",
            "/html/body/form/div[2]/div[33]/table/tbody/tr[1]/td/label",
            "/html/body/form/div[2]/div[35]/table/tbody/tr[1]/td/label"
        ]
        
        for checkbox_xpath in checkbox_xpaths:
            try:
                checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
                driver.execute_script("arguments[0].click();", checkbox)
                print(f"Checked {checkbox_xpath}")
                time.sleep(1)  # Add small delay between checkbox clicks
            except Exception as e:
                print(f"Error checking checkbox {checkbox_xpath}: {str(e)}")
                driver.save_screenshot(f"checkbox_error_{checkbox_xpath.replace('/', '_')}.png")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")
        
        # Submit form
        submit_xpath = "/html/body/form/div[3]/div/input"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    


def handle_claytoncountywaterauthority_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/form/ul/li[1]/span[1]/input": form_data["first name"],
            "/html/body/div[1]/form/ul/li[1]/span[2]/input": form_data["last name"],
            "/html/body/div[1]/form/ul/li[5]/span[1]/input": form_data["p1"],
            "/html/body/div[1]/form/ul/li[5]/span[3]/input": form_data["p2"],
            "/html/body/div[1]/form/ul/li[5]/span[5]/input": form_data["p3"],
            "/html/body/div[1]/form/ul/li[6]/div/input": form_data["email"],
            "/html/body/div[1]/form/ul/li[4]/div/input": form_data["address"],
          
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/form/ul/li[7]/div/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/form/ul/li[8]/div/input[2]"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_peachtreecitygapolice_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
        
            
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[1]/div/input": form_data["email"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[2]/div/input": form_data["name"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[3]/div/input": form_data["phone"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[4]/div/textarea": form_data["address"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[1]/div/div/input": form_data["city"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[3]/div/div/input": form_data["zip"],
            "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[6]/div/input": form_data["company"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[1]/div[2]/div[1]/div[2]/div/div[1]"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[2]/div[5]/div[2]/div/div/div/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Alabama")  # Adjust text as per available options
        print("Selected dropdown option")

                    # Handle dropdown selection (corrected)
        dropdown2_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[1]/div[4]/div/div/div/div/div[1]/input"
        dropdown2 = wait.until(EC.presence_of_element_located((By.XPATH, dropdown2_xpath)))



   # Clear the dropdown input field
        dropdown2.clear()

# Enter the desired option into the input field
        dropdown2.send_keys("Library")
        print("Typed 'Library' into the dropdown")

# Press Enter to confirm the selection
        dropdown2.send_keys(Keys.ENTER)
        print("Pressed Enter to select 'Library'")
        # check
        checkbox1_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[3]/div[2]/div/div/input"
        for checkbox_xpath in [checkbox1_xpath]:
            checkbox = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
            driver.execute_script("arguments[0].click();", checkbox)
            print(f"Checked {checkbox_xpath}")
        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/main/div/main/section/div/div[1]/form/div[4]/button"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")
        time.sleep(20)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your request reference number is')]").text
            print(f"Confirmation message: {confirmation_message}")
            security_key_match = re.search(r"Your security key is (\S+)", confirmation_message)
            reference_number_match = re.search(r"Your request reference number is (\S+)", confirmation_message)
            security_key = security_key_match.group(1) if security_key_match else ""
            reference_number = reference_number_match.group(1) if reference_number_match else ""
            if not security_key or not reference_number:
                print("Confirmation message found but could not extract reference number or security key")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    

def handle_riverdale_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}


def handle_forestparkga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[10]/div/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[11]/div/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(40)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}
    
def handle_acworthga_form(driver, url):
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        # Wait for page to load completely
        time.sleep(7)

        # Fill all fields
        fields_to_fill = {
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[1]/div[1]/input": form_data["name"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[2]/div[1]/input": form_data["phone"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[3]/div[1]/input": form_data["email"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[4]/div[1]/input": form_data["address"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[5]/div[1]/input": form_data["city"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[6]/div[1]/input": form_data["state"],
            "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[7]/div[1]/input": form_data["zip"]
        }

        for xpath, value in fields_to_fill.items():
            fill_field(driver, wait, xpath, value)

        # Fill request details
        details_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[8]/div[1]/textarea"
        details = wait.until(EC.presence_of_element_located((By.XPATH, details_xpath)))
        details.clear()
        details.send_keys(form_data["message"])
        print("Filled request details")

        # Handle dropdown selection
        dropdown1_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[2]/div/div[9]/div[1]/select"
        dropdown1 = Select(wait.until(EC.presence_of_element_located((By.XPATH, dropdown1_xpath))))
        available_options = [o.text.strip() for o in dropdown1.options]
        print("Available options in dropdown:", available_options)
        dropdown1.select_by_visible_text("Only to review / inspect")  # Adjust text as per available options
        print("Selected dropdown option")

        # Take screenshot before submission
        driver.save_screenshot("before_submit.png")

        # Submit form
        submit_xpath = "/html/body/div[1]/div[2]/div[1]/form/main/div/div[4]/div/button/div"
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        driver.execute_script("arguments[0].click();", submit_button)
        print("Clicked submit button")

        # Delay of 15 seconds after form submission
        time.sleep(60)

        # Extract confirmation message
        try:
            confirmation_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Your security key is')]").text
            security_key = re.search(r"Your security key is (\S+)", confirmation_message).group(1)
            reference_number = re.search(r"Your request reference number is (\S+)", confirmation_message).group(1)
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": reference_number, "security_key": security_key}
        except NoSuchElementException:
            print("Confirmation message not found")
            return {"status": "Success", "confirmation": "Form submitted", "reference_number": "", "security_key": ""}

    except Exception as e:
        print(f"Error in form handling: {str(e)}")
        driver.save_screenshot("error_main.png")
        return {"status": "Failed", "error": str(e)}



def process_form(url,retries=1):
    for attempt in range(retries):
        driver = create_driver()
        wait = WebDriverWait(driver, 20)
        driver.get(url)
        try:
            if url == "https://cantonga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_cantonga_form(driver, url)
            elif url == "https://maconbibbcountyga.justfoia.com/Forms/Launch/a709d888-de2f-4857-9c82-b12b2645a87c":
                result = handle_maconbibbcountyga_form(driver, url)
            elif url == "https://woodstockga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_woodstockga_form(driver, url)
            elif url == "https://smyrnaga.justfoia.com/Forms/Launch/fd208f47-7557-4edf-9478-723c87ba6b30":
                result = handle_smyrnaga_form(driver, url)
            elif url == "https://austellga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_austellga_form(driver, url)
            elif url == "https://albanyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_albanyga_form(driver, url)
            elif url == "https://forsythcountyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_forsythcountyga_form(driver, url)
            elif url == "https://fairburnga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_fairburnga_form(driver, url)
            elif url == "https://collegeparkga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_collegeparkga_form(driver, url)
            elif url == "https://conyersga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_conyersga_form(driver, url)
            elif url == "https://spaldingcountyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_spaldingcountyga_form(driver, url)
            elif url == "https://cityofaugustaga.nextrequest.com/requests/new":
                result = handle_cityofaugustaga_form(driver, url)
            elif url == "https://cityoffayettevillega.nextrequest.com/requests/new":
                result = handle_cityoffayettevillega_form(driver, url)
            elif url == "https://unioncityga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_unioncityga_form(driver, url)
            elif url == "https://hapevillega.justfoia.com/Forms/Launch/0e8fe236-181b-486e-8bb2-33990d9295c0":
                result = handle_hapevillega_form(driver, url)
            elif url == "https://eastpointga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_eastpointga_form(driver, url)
            elif url == "https://sandyspringsga.justfoia.com/Forms/Launch/9739a52c-6627-4522-acf8-5f87f95f4a9d":
                result = handle_sandyspringsga_form(driver, url)
            elif url == "https://roswellga.justfoia.com/Forms/Launch/02b6cc8e-5783-45b0-9fbb-3779b3387e72":
                result = handle_roswellga_form(driver, url)
            elif url == "https://alpharettaga.justfoia.com/Forms/Launch/39f66254-4f9f-4794-955f-a15b1afebd0e":
                result = handle_alpharettaga_form(driver, url)
            elif url == "https://norcrossga.justfoia.com/Forms/Launch/2da76d30-7849-4d56-b182-ef68865e40f6":
                result = handle_norcrossga_form(driver, url)
            elif url == "https://henrycounty-services.app.transform.civicplus.com/forms/34175":
                result = handle_henrycounty_form(driver, url)
            elif url == "https://fayettecountyga.gov/administration/open-records-request":
                result = handle_fayettecountyga_form(driver, url)
            elif url == "https://fs6.formsite.com/mAFRD/jiupubq3at/index.html":
                result = handle_formsite_form(driver, url)
            elif url == "https://claytoncountywaterauthority.wufoo.com/forms/z1e6l46517vub7j/":
                result = handle_claytoncountywaterauthority_form(driver, url)
            elif url == "https://peachtreecitygapolice.nextrequest.com/requests/new":
                result = handle_peachtreecitygapolice_form(driver, url)
            elif url == "https://riverdalega.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_riverdale_form(driver, url)
            elif url == "https://forestparkga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_forestparkga_form(driver, url)
            elif url == "https://acworthga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb":
                result = handle_acworthga_form(driver, url)
            else:
                result = {"status": "Success", "confirmation": "Form submitted", "reference_number": "N/A", "security_key": "N/A", "URL": url}
        except Exception as e:
            print(f"Attempt {attempt + 1} failed for {url}: {str(e)}")
            result = {"status": "failed", "confirmation": "", "error": str(e), "URL": url}
            time.sleep(60) 
        finally:
            driver.quit()
    
    return result

def save_results(county_name, results, save_path, filename):
    # Define the path to the spreadsheet
    START_DATE = datetime.now().strftime("%Y-%m-%d")
    file_name = START_DATE + ".xlsx"
    file_name = file_name.replace('/', '-')  # Ensure proper filename format
    file_path = os.path.join(save_path, file_name)
    print(f"Saving results for {county_name} to {file_path}")

    # Check if the spreadsheet already exists
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    # Remove the default sheet if it exists
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Create or select the Responses sheet
    if "Responses" in workbook.sheetnames:
        responses_sheet = workbook["Responses"]
    else:
        responses_sheet = workbook.create_sheet(title="Responses")

    # Define the yellow fill for header cells
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define the headers for the Responses sheet
    responses_headers = [
        'Date Of Request', 'County', 'URL', 'Status', 'Reference Number', 'Security Key', 'Notes'
    ]

    # Write headers if the Responses sheet is new
    if responses_sheet.max_row == 1:
        responses_sheet.append(responses_headers)
        for col_num, header in enumerate(responses_headers, 1):
            cell = responses_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = yellow_fill
        responses_sheet.row_dimensions[1].height = 30  # Increase the height of the header row

    # Append the results to the Responses sheet
    for result in results:
        row = [
            datetime.now().strftime("%Y-%m-%d"),  # Date Of Request
            result.get('county', ""),  # County
            result.get('url', ""),  # URL
            result.get('status', ""),  # Status
            result.get('reference_number', ""),  # Reference Number
            result.get('security_key', ""),  # Security Key
            result.get('Notes', "")  # Notes
        ]
        responses_sheet.append(row)

    # Adjust column widths to fit content for the Responses sheet
    for column in responses_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 5  # Add padding for better view
        responses_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    try:
        # Ensure file is saved with appropriate encoding
        safe_file_path = os.path.join(save_path, filename)
        workbook.save(safe_file_path)
        print(f"Results successfully saved to {safe_file_path}")
    except PermissionError:
        print(f"Permission denied: Unable to save the file at {safe_file_path}. Please close the file if it is open and try again.")
    except Exception as e:
        print(f"An error occurred while saving the file: {str(e)}")
        
def update_progress_bar(progress_bar, progress_var, value, progress_label, county):
    progress_var.set(value)
    progress_label.config(text=f"Processing {county}...")
    progress_bar.update_idletasks()


def run_processing(urls, results, progress_queue, output_filename, save_path, county_name):
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_url = {executor.submit(process_form, url): url for url in urls}
        for i, future in enumerate(concurrent.futures.as_completed(future_to_url)):
            url = future_to_url[future]
            try:
                result = future.result()
                results.append({
                    "county": county_name,
                    "url": url,
                    "status": result["status"],
                    "confirmation": result.get("confirmation", ""),
                    "error": result.get("error", ""),
                    "reference_number": result.get("reference_number", ""),
                    "security_key": result.get("security_key", "")
                })
            except Exception as e:
                results.append({
                    "county": county_name,
                    "url": url,
                    "status": "failed",
                    "confirmation": "",
                    "error": str(e),
                    "reference_number": "",
                    "security_key": ""
                })
            progress_queue.put((i + 1, f"Processing {url}"))

    # Save results
    print(results)
    save_results(county_name, results, save_path, output_filename)

    # Signal that processing is complete
    progress_queue.put((None, "Processing complete"))

def create_gradient(canvas, width, height):
    for i in range(height):
        grey_value = int(200 - (i * 100 / height))
        canvas.create_rectangle(0, 0, width, height, fill='grey', outline='grey')
        color = f'#{grey_value:02x}{grey_value:02x}{grey_value:02x}'  # Define the color as a shade of grey
        canvas.create_line(0, i, width, i, fill=color)

def resize_canvas(event):
    create_gradient(event.widget, event.width, event.height)

def submit_form():
    global form_data
    form_data = {
        "date": date_entry.get(),
        "name": name_entry.get(),
        "first name": first_name_entry.get(),
        "last name": last_name_entry.get(),
        "phone": phone_entry.get(),
        "email": email_entry.get(),
        "address": address_entry.get(),
        "city": city_entry.get(),
        "state": state_entry.get(),
        "zip": zip_entry.get(),
        "company": company_entry.get(),
        "case": case_entry.get(),
        "time": time_entry.get(),
        "person represented": person_represented_entry.get(),
        "case number": case_number_entry.get(),
        "unit number": unit_number_entry.get(),
        "country": country_entry.get(),
        "message": message_entry.get("1.0", tk.END),
        "output path": entries.get("output_path_entry", ""),
        "output filename": output_filename_entry.get()
    }
    
    # Check if all fields are filled
    for key, value in form_data.items():
        if not value.strip() and key not in ["output path", "output filename"]:
            messagebox.showerror("Error", f"Please fill in the {key} field.")
            return
    
    selected_counties = [county for county, var in county_vars.items() if var.get()]
    if not selected_counties:
        messagebox.showerror("Error", "Please select at least one county.")
        return
    
    # Withdraw the form window
    root.withdraw()
    
    # Create a new window for the progress bar
    global progress_window, progress_var, progress_bar, progress_queue, progress_label
    progress_window = tk.Toplevel(root)
    progress_window.title("Processing Progress")
    
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=len(selected_counties), length=400)
    progress_bar.pack(pady=20, padx=20)
    
    progress_label = tk.Label(progress_window, text="Started processing...")
    progress_label.pack(pady=10)
    
    progress_queue = queue.Queue()
    
    def process_queue():
        try:
            value, message = progress_queue.get_nowait()
            if value is None:
                progress_window.quit()
            else:
                update_progress_bar(progress_bar, progress_var, value, progress_label, message)
            progress_window.after(100, process_queue)
        except queue.Empty:
            progress_window.after(100, process_queue)
    
    process_queue()
    
    output_filename = f"{form_data['output path']}/{form_data['output filename']}.csv"
    save_path = form_data['output path']
    
    for county in selected_counties:
        if county in urls:
            threading.Thread(target=run_processing, args=(urls[county], results, progress_queue, output_filename, save_path, county)).start()
def draw_form_fields():
    global entries
    entries = {}
    x_position = 50
    y_position = 50
    for i, (field, var_name) in enumerate([
        ("Date", "date_entry"),
        ("Name", "name_entry"),
        ("First Name", "first_name_entry"),
        ("Last Name", "last_name_entry"),
        ("Phone", "phone_entry"),
        ("Email", "email_entry"),
        ("Address", "address_entry"),
        ("City", "city_entry"),
        ("State", "state_entry"),
        ("Zip", "zip_entry"),
        ("Company", "company_entry"),
        ("Case", "case_entry"),
        ("Time", "time_entry"),
        ("Person Represented", "person_represented_entry"),
        ("Case Number", "case_number_entry"),
        ("Unit Number", "unit_number_entry"),
        ("Country", "country_entry"),
        ("Message", "message_entry"),
        ("Output Filename", "output_filename_entry")
    ]):
        label = ttk.Label(form_frame, text=field, font=("Helvetica", 10, "bold"), background="#808080") 
        canvas.create_window(x_position, y_position, window=label, anchor="w")
        if field == "Message":
            entry = tk.Text(form_frame, height=5, width=50)
        else:
            entry = ttk.Entry(form_frame)
        canvas.create_window(x_position + 200, y_position, window=entry, anchor="w")
        entries[var_name] = entry

        if (i + 1) % 3 == 0:
            x_position = 50
            y_position += 70 if field != "Message" else 100
        else:
            x_position += 400

    global date_entry, name_entry, first_name_entry, last_name_entry, phone_entry, email_entry, address_entry, city_entry, state_entry, zip_entry, company_entry, case_entry, time_entry, person_represented_entry, case_number_entry, unit_number_entry, country_entry, message_entry, output_filename_entry
    date_entry = entries["date_entry"]
    name_entry = entries["name_entry"]
    first_name_entry = entries["first_name_entry"]
    last_name_entry = entries["last_name_entry"]
    phone_entry = entries["phone_entry"]
    email_entry = entries["email_entry"]
    address_entry = entries["address_entry"]
    city_entry = entries["city_entry"]
    state_entry = entries["state_entry"]
    zip_entry = entries["zip_entry"]
    company_entry = entries["company_entry"]
    case_entry = entries["case_entry"]
    time_entry = entries["time_entry"]
    person_represented_entry = entries["person_represented_entry"]
    case_number_entry = entries["case_number_entry"]
    unit_number_entry = entries["unit_number_entry"]
    country_entry = entries["country_entry"]
    message_entry = entries["message_entry"]
    output_filename_entry = entries["output_filename_entry"]

    def select_directory_dialog():
        directory_path = filedialog.askdirectory()
        if directory_path:
            entries["output_path_entry"] = directory_path
        root.lift()

    def save_template():
        template_data = {var_name: entry.get("1.0", tk.END).strip() if isinstance(entry, tk.Text) else entry.get() for var_name, entry in entries.items()}
        save_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if save_path:
            with open(save_path, 'w') as json_file:
                json.dump(template_data, json_file, indent=4)
            messagebox.showinfo("Success", "Template saved successfully.")
        root.lift()

    directory_button = ttk.Button(form_frame, text="Select Directory", command=select_directory_dialog)
    canvas.create_window(300, y_position + 50, window=directory_button, anchor="w")

    save_template_button = ttk.Button(form_frame, text="Save Template", command=save_template)
    canvas.create_window(500, y_position + 50, window=save_template_button, anchor="w")

    submit_button = ttk.Button(form_frame, text="Submit", command=submit_form)
    canvas.create_window(300, y_position + 100, window=submit_button, anchor="w")

def open_form_window():
    global root, canvas, form_frame, json_frame
    root = tk.Toplevel()
    root.title("Form Submission Progress")

    # Create a notebook for tabs
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)

    # Create a frame for the form
    form_frame = ttk.Frame(notebook)
    notebook.add(form_frame, text="Manual Form")

    # Create a frame for JSON file selection
    json_frame = ttk.Frame(notebook)
    notebook.add(json_frame, text="Upload JSON")

    # Create a canvas for the form
    global canvas
    canvas = tk.Canvas(form_frame, width=1000, height=1000, bg="#808080")
    canvas.pack(fill="both", expand=True)

    # Create the gradient background
    create_gradient(canvas, 1000, 1000)
    canvas.bind("<Configure>", resize_canvas)

    draw_form_fields()

    # Create the gradient background for JSON frame
    json_canvas = tk.Canvas(json_frame, width=1000, height=1000, bg="#808080")
    json_canvas.pack(fill="both", expand=True)
    create_gradient(json_canvas, 1000, 1000)
    json_canvas.bind("<Configure>", resize_canvas)

    # Add JSON file selection components
    json_label = ttk.Label(json_canvas, text="Select JSON file to pre-fill the form:", font=("Helvetica", 10, "bold"), background="#808080")
    json_canvas.create_window(500, 100, window=json_label)
    json_button = ttk.Button(json_canvas, text="Browse", command=load_json_file)
    json_canvas.create_window(500, 150, window=json_button)

def load_json_file():
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if file_path:
        with open(file_path, 'r') as file:
            try:
                data = json.load(file)
                fill_form_with_json(data)
            except json.JSONDecodeError:
                messagebox.showerror("Error", "Invalid JSON file.")
        root.lift()

def fill_form_with_json(data):
    try:
        date_entry.insert(0, data["date"])
        name_entry.insert(0, data["name"])
        first_name_entry.insert(0, data["first name"])
        last_name_entry.insert(0, data["last name"])
        phone_entry.insert(0, data["phone"])
        email_entry.insert(0, data["email"])
        address_entry.insert(0, data["address"])
        city_entry.insert(0, data["city"])
        state_entry.insert(0, data["state"])
        zip_entry.insert(0, data["zip"])
        company_entry.insert(0, data["company"])
        case_entry.insert(0, data["case"])
        time_entry.insert(0, data["time"])
        person_represented_entry.insert(0, data["person represented"])
        case_number_entry.insert(0, data["case number"])
        unit_number_entry.insert(0, data["unit number"])
        country_entry.insert(0, data["country"])
        message_entry.insert("1.0", data["message"])
    except KeyError as e:
        messagebox.showerror("Error", f"Missing key in JSON file: {e}")
    root.lift()
    
def select_counties():
    selected_counties = [county for county, var in county_vars.items() if var.get()]
    if not selected_counties:
        messagebox.showerror("Error", "Please select at least one county.")
        return
    county_selection_window.destroy()
    open_form_window()

def draw_county_selection():
    global county_vars, county_selection_window
    county_selection_window = tk.Tk()
    county_selection_window.title("Select Counties")

    canvas = tk.Canvas(county_selection_window, width=1000, height=500)
    canvas.pack(fill="both", expand=True)

    # Create the gradient background
    create_gradient(canvas, 1000, 500)
    canvas.bind("<Configure>", resize_canvas)

    x_position = 50
    y_position = 50
    county_vars = {}

    # Add Select All checkbox
    select_all_var = tk.BooleanVar()
    select_all_checkbox = ttk.Checkbutton(county_selection_window, text="Select All", variable=select_all_var, style="County.TCheckbutton")
    canvas.create_window(x_position, y_position, window=select_all_checkbox, anchor="w")
    y_position += 30

    def toggle_select_all():
        for var in county_vars.values():
            var.set(select_all_var.get())

    select_all_var.trace_add("write", lambda *args: toggle_select_all())

    for county in urls.keys():
        var = tk.BooleanVar()
        checkbox = ttk.Checkbutton(county_selection_window, text=county, variable=var, style="County.TCheckbutton")
        canvas.create_window(x_position, y_position, window=checkbox, anchor="w")
        county_vars[county] = var
        x_position += 200
        if x_position > 800:
            x_position = 50
            y_position += 30

    def on_next():
        selected_counties = [county for county, var in county_vars.items() if var.get()]
        if not selected_counties:
            messagebox.showerror("Error", "Please select at least one county.")
            return
        county_selection_window.withdraw()  # Hide the county selection window
        open_form_window()

    next_button = ttk.Button(county_selection_window, text="Next", command=on_next)
    canvas.create_window(450, y_position + 50, window=next_button, anchor="w")

    # Style for checkbuttons
    style = ttk.Style()
    style.configure("County.TCheckbutton", font=("Helvetica", 10, "bold"), background="#808080")

    county_selection_window.mainloop()
    
if __name__ == "__main__":
    results = []

    # Dictionary of URLs classified by county
    urls = {
        "Bibb County": [
            "https://maconbibbcountyga.justfoia.com/Forms/Launch/a709d888-de2f-4857-9c82-b12b2645a87c"
        ],
        "Cherokee County": [
            "https://cherokeecountyga.nextrequest.com/requests/new",
            "https://cantonga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://woodstockga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ],
        "Clayton County": [
            "https://riverdalega.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://forestparkga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://claytoncountywaterauthority.wufoo.com/forms/z1e6l46517vub7j/"
        ],
        "Cobb County": [
            "https://austellga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://acworthga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://kennesawga.justfoia.com/Forms/Launch/a3570e65-d44d-43d3-a822-d38e2fc1c3d3",
            "https://smyrnaga.justfoia.com/Forms/Launch/fd208f47-7557-4edf-9478-723c87ba6b30"
        ],
        "Dekalb County": [
            "https://fs6.formsite.com/mAFRD/jiupubq3at/index.html"
        ],
        "Dougherty County": [
            "https://albanyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ],
        "Fayette County": [
            "https://fayettecountyga.gov/administration/open-records-request",
            "https://cityoffayettevillega.nextrequest.com/requests/new",
            "https://peachtreecitygapolice.nextrequest.com/requests/new"
        ],
        "Forsyth County": [
            "https://forsythcountyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ],
        "Fulton County": [
            "https://fs6.formsite.com/mAFRD/jiupubq3at/index.html",
            "https://alpharettaga.justfoia.com/Forms/Launch/39f66254-4f9f-4794-955f-a15b1afebd0e",
            "https://roswellga.justfoia.com/Forms/Launch/02b6cc8e-5783-45b0-9fbb-3779b3387e72",
            "https://fairburnga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://sandyspringsga.justfoia.com/Forms/Launch/9739a52c-6627-4522-acf8-5f87f95f4a9d",
            "https://eastpointga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://unioncityga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb",
            "https://hapevillega.justfoia.com/Forms/Launch/0e8fe236-181b-486e-8bb2-33990d9295c0",
            "https://collegeparkga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ],
        "Gwinnett County": [
            "https://norcrossga.justfoia.com/Forms/Launch/2da76d30-7849-4d56-b182-ef68865e40f6"
        ],
        "Henry County": [
            "https://henrycounty-services.app.transform.civicplus.com/forms/34175"
        ],
        "Richmond County": [
            "https://cityofaugustaga.nextrequest.com/requests/new"
        ],
        "Rockdale County": [
            "https://conyersga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ],
        "Spalding County": [
            "https://spaldingcountyga.justfoia.com/Forms/Launch/d705cbd6-1396-49b7-939e-8d86c5a87deb"
        ]
    }
    
    draw_county_selection()


